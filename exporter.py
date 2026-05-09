"""Export results to a color-coded Excel report."""
from __future__ import annotations
from pathlib import Path
from typing import List, Union, Iterable
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .comparator import CompareResult

# Colors (Excel-style soft fills)
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

INVALID_SHEETNAME_CHARS = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_name(name: str, used: Iterable[str]) -> str:
    cleaned = INVALID_SHEETNAME_CHARS.sub("_", name).strip()[:31] or "Sheet"
    used_set = set(used)
    if cleaned not in used_set:
        return cleaned
    # disambiguate
    base = cleaned[:28]
    for i in range(1, 1000):
        candidate = f"{base}_{i}"[:31]
        if candidate not in used_set:
            return candidate
    return cleaned


def _write_df(wb: Workbook, name: str, df: pd.DataFrame, fill: PatternFill = None):
    name = _safe_sheet_name(name, wb.sheetnames)
    ws = wb.create_sheet(title=name)
    if df is None or df.empty:
        ws.append(["(no rows)"])
        return
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows):
        ws.append(row)
        if r_idx == 0:
            for cell in ws[1]:
                cell.fill = HEADER
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
        elif fill is not None:
            for cell in ws[r_idx + 1]:
                cell.fill = fill
    # Auto-width-ish (cap at 40)
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


def export_results(
    results: List[CompareResult],
    output_path: Union[str, Path],
) -> Path:
    """Write all results into one XLSX with a Summary sheet first."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    summary_df = pd.DataFrame([r.summary for r in results])
    _write_df(wb, "Summary", summary_df)

    # Per-result detail sheets
    for i, r in enumerate(results, 1):
        stem = Path(r.source).stem if r.source else f"result{i}"
        prefix = f"{i:02d}_{stem[:12]}"
        sheet_tag = f"_{r.sheet[:6]}" if r.sheet and r.sheet != "Sheet1" else ""
        _write_df(wb, f"{prefix}{sheet_tag}_match", r.matched, GREEN)
        _write_df(wb, f"{prefix}{sheet_tag}_unmatch", r.unmatched, RED)
        if r.diffs is not None and not r.diffs.empty:
            _write_df(wb, f"{prefix}{sheet_tag}_diff", r.diffs, AMBER)

    wb.save(output)
    return output

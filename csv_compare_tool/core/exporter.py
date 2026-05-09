"""Export results to a color-coded Excel report."""
from __future__ import annotations
import io
from pathlib import Path
from typing import List, Union, Iterable
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

from .comparator import CompareResult

# Colors (Excel-style soft fills)
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

INVALID_SHEETNAME_CHARS = re.compile(r'[\\/*?:\[\]]')


def _safe_sheet_name(name: str, used: Iterable[str]) -> str:
    cleaned = INVALID_SHEETNAME_CHARS.sub("_", name).strip()[:31] or "Sheet"
    used_set = set(used)
    if cleaned not in used_set:
        return cleaned
    # disambiguate
    base = cleaned[:28]
    for i in range(1, 1000):
        candidate = f"{base}_{i}"[:31]
        if candidate not in used_set:
            return candidate
    return cleaned


def _write_df(wb: Workbook, name: str, df: pd.DataFrame, fill: PatternFill = None):
    name = _safe_sheet_name(name, wb.sheetnames)
    ws = wb.create_sheet(title=name)
    if df is None or df.empty:
        ws.append(["(no rows)"])
        return
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows):
        ws.append(row)
        if r_idx == 0:
            for cell in ws[1]:
                cell.fill = HEADER
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
        elif fill is not None:
            for cell in ws[r_idx + 1]:
                cell.fill = fill
    # Auto-width-ish (cap at 40)
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)


def export_results(
    results: List[CompareResult],
    output_path: Union[str, Path],
) -> Path:
    """Write all results into one XLSX with a Summary sheet first."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    summary_df = pd.DataFrame([r.summary for r in results])
    _write_df(wb, "Summary", summary_df)

    # Per-result detail sheets
    for i, r in enumerate(results, 1):
        stem = Path(r.source).stem if r.source else f"result{i}"
        prefix = f"{i:02d}_{stem[:12]}"
        sheet_tag = f"_{r.sheet[:6]}" if r.sheet and r.sheet != "Sheet1" else ""
        _write_df(wb, f"{prefix}{sheet_tag}_match", r.matched, GREEN)
        _write_df(wb, f"{prefix}{sheet_tag}_unmatch", r.unmatched, RED)
        if r.diffs is not None and not r.diffs.empty:
            _write_df(wb, f"{prefix}{sheet_tag}_diff", r.diffs, AMBER)

    wb.save(output)
    return output


def export_pci_results(results: list, output_path: Union[str, Path]) -> Path:
    """Write PCI analysis results to XLSX: Summary + per-file detail + embedded pie charts."""
    from matplotlib.figure import Figure
    from matplotlib.backends.backend_agg import FigureCanvasAgg

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ────────────────────────────────────────────
    summary_df = pd.DataFrame([r.summary for r in results])
    _write_df(wb, "Summary", summary_df)

    # ── Per-file detail sheets ────────────────────────────────────
    for idx, r in enumerate(results, 1):
        # Sheet name = FloorName_ShortMeasurementName (Excel limit: 31 chars)
        floor = r.floor[:12] if r.floor else f"{idx:02d}"
        short = r.short_name[:16] if r.short_name else "detail"
        sheet_name = _safe_sheet_name(f"{floor}_{short}", wb.sheetnames)
        ws = wb.create_sheet(title=sheet_name)

        detail_df = r.detail_df()
        rows = list(dataframe_to_rows(detail_df, index=False, header=True))
        for r_idx, row_data in enumerate(rows):
            ws.append(row_data)
            if r_idx == 0:
                for cell in ws[1]:
                    cell.fill = HEADER
                    cell.font = HEADER_FONT
                    cell.alignment = HEADER_ALIGN
            else:
                in_ref = len(row_data) > 3 and str(row_data[3]) == "Yes"
                fill = GREEN if in_ref else RED
                for cell in ws[r_idx + 1]:
                    cell.fill = fill

        for col_idx, col in enumerate(detail_df.columns, 1):
            max_len = max(
                [len(str(col))] + [len(str(v)) for v in detail_df[col].head(200).tolist()]
            )
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = min(max_len + 2, 40)

        # ── Embedded pie chart ────────────────────────────────────
        try:
            counts = r.pci_counts
            if not counts.empty:
                max_slices = 15
                if len(counts) > max_slices:
                    top    = counts.iloc[:max_slices]
                    rest   = counts.iloc[max_slices:]
                    labels = list(top.index)
                    values = list(top.values)
                    colors = ["#70AD47" if p in r.in_ref_pcis else "#FF4444" for p in top.index]
                    rest_r = int(rest[rest.index.isin(r.in_ref_pcis)].sum())
                    rest_u = int(rest[rest.index.isin(r.unique_pcis)].sum())
                    if rest_r:
                        labels.append("Others (Ref)"); values.append(rest_r); colors.append("#A9D18E")
                    if rest_u:
                        labels.append("Others (Unique)"); values.append(rest_u); colors.append("#FF9999")
                else:
                    labels = list(counts.index)
                    values = list(counts.values)
                    colors = ["#70AD47" if p in r.in_ref_pcis else "#FF4444" for p in counts.index]

                total = sum(values)
                fig = Figure(figsize=(7, 5.5), constrained_layout=True)
                FigureCanvasAgg(fig)
                ax = fig.add_subplot(111)
                wedges, _, autotexts = ax.pie(
                    values, colors=colors,
                    autopct=lambda pct: f"{pct:.1f}%" if pct >= 3 else "",
                    startangle=90,
                    wedgeprops={"linewidth": 0.5, "edgecolor": "white"},
                )
                for at in autotexts:
                    at.set_fontsize(8)
                legend_labels = [
                    f"PCI {lbl}  —  {v:,} ({v/total*100:.1f}%)"
                    for lbl, v in zip(labels, values)
                ]
                ax.legend(wedges, legend_labels,
                          loc="lower center", bbox_to_anchor=(0.5, -0.28),
                          ncol=min(2, max(1, len(labels) // 7 + 1)),
                          fontsize=7)
                chart_title = f"[{r.floor}]  {r.short_name}" if r.floor else r.short_name
                if len(chart_title) > 60:
                    chart_title = chart_title[:58] + "…"
                ax.set_title(
                    f"{chart_title}\n"
                    f"{r.total_rows:,} rows  |  "
                    f"{len(r.in_ref_pcis)} in ref  |  {len(r.unique_pcis)} unique",
                    fontsize=9,
                )
                buf = io.BytesIO()
                fig.savefig(buf, format="png", dpi=110)
                buf.seek(0)
                xl_img = XLImage(buf)
                xl_img.anchor = "F2"
                ws.add_image(xl_img)
        except Exception:
            pass  # chart is optional

    wb.save(output)
    return output
